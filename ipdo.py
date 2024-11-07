import pandas as pd 
import os 
import shutil 
from datetime import datetime, timedelta
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import openpyxl
import time 
import warnings
import matplotlib.pyplot as plt
import seaborn as sns

warnings.filterwarnings("ignore")
# Configurações do Seaborn para melhorar a aparência dos gráficos
sns.set_theme(style="darkgrid")

def get_info():
    informacoes = {}
    
    # Abre o arquivo e lê linha por linha
    with open('config.txt', 'r', encoding='utf-8') as arquivo:
        
        for linha in arquivo:
            # Remove espaços em branco nas bordas e quebra de linha
            linha = linha.strip()
            
            # Divide a linha no formato chave:valor
            if ":" in linha:
                chave, valor = linha.split(":", 1)
                chave = chave.strip()
                valor = valor.strip()
                
                # Armazena no dicionário
                informacoes[chave] = valor
    
    return informacoes

def ipdo_download():
    servico = Service(ChromeDriverManager().install())
    # Instancia o Chrome options
    chrome_options = webdriver.ChromeOptions()
    # Ativar modo headless
    # chrome_options.add_argument("--headless=new")  
    # chrome_options.add_argument("--no-sandbox")
    # chrome_options.add_argument("--disable-dev-shm-usage")

    # Inicializando o driver do Chrome no modo headless
    driver = webdriver.Chrome(service=servico, options=chrome_options)

    # Tempo de espera do navegador para encontrar um elemento
    wait_driver = WebDriverWait(driver, 10)

    driver.get('https://sintegre.ons.org.br/sites/7/39/Paginas/servicos/historico-de-produtos.aspx?produto=IPDO%20Edit%C3%A1vel')
    wait_driver.until(EC.element_to_be_clickable((By.ID, 'username')))
    driver.find_element(By.ID,'username').send_keys(config['Usuário sintegre'])
    wait_driver.until(EC.element_to_be_clickable((By.ID, 'password')))
    driver.find_element(By.ID,'password').send_keys(config['Senha sintegre'])
    driver.find_element(By.ID, 'kc-login').click()
    # Retirar Mensagem Cookies 
    wait_driver.until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/div[12]/div/div[5]/button')))
    driver.find_element(By.XPATH, '/html/body/form/div[12]/div/div[5]/button').click()

    for i in range (1,11): 
        try:
            wait_driver.until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/div[12]/div/div[3]/div/div/div[2]/div[2]/div/div/div/div[2]/div[1]/div/div/div/div[2]/div/div[2]/div[2]/div[3]/a[2]')))
        except:
            driver.refresh()
            wait_driver.until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/div[12]/div/div[3]/div/div/div[2]/div[2]/div/div/div/div[2]/div[1]/div/div/div/div[2]/div/div[2]/div[2]/div[3]/a[2]')))
        nome_arquivo_ipdo = driver.find_element(By.XPATH,f'/html/body/form/div[12]/div/div[3]/div/div/div[2]/div[2]/div/div/div/div[2]/div[1]/div/div/div/div[2]/div/div[{i}]/div[2]/div[2]/a').text
        download_arquivo_ipdo = driver.find_element(By.XPATH,f'/html/body/form/div[12]/div/div[3]/div/div/div[2]/div[2]/div/div/div/div[2]/div[1]/div/div/div/div[2]/div/div[{i}]/div[2]/div[3]/a[2]')
        if os.path.exists(os.path.join('ipdo', nome_arquivo_ipdo)):
            pass
        else:
            if nome_arquivo_ipdo.endswith('.xlsm'):
                download_arquivo_ipdo.click()
                time.sleep(10)
                shutil.move(f'{config["Diretório download"]}{nome_arquivo_ipdo}',f'ipdo/{nome_arquivo_ipdo}')
    driver.close()

def def_datas():
    dia_atual = datetime.today()    
    
    data = pd.to_datetime(config['Data'])

    lista_datas = []
    intervalo = timedelta(days=1)
    while data <= dia_atual:
        lista_datas.append(data)
        data += intervalo
    lista_datas_str = [data.strftime("%d-%m-%Y") for data in lista_datas][:-1]

    return lista_datas_str

def create_df():
    df_ipdo = pd.DataFrame()

    for arquivo in datas_ipdo:
        data_dia_ipdo = arquivo
        # Monta o nome do arquivo
        arquivo = 'IPDO-'+arquivo+'.xlsm'
        try:
        # Abre o arquivo como workbook
            workbook = openpyxl.load_workbook(f'ipdo/{arquivo}', read_only=True)
        except():
            pass
        # Abre a planilha IPDO
        sheet = workbook['IPDO']
        # Requisita os dados nas células específicas
        dados = {'Data':data_dia_ipdo,'ENA Sudeste': (sheet['M65'].value),'ENA Sul': (sheet['M64'].value),'ENA Norte': (sheet['M62'].value),'ENA Nordeste': (sheet['M63'].value),'MLT Sudeste': (sheet['O65'].value)*100,'MLT Sul': (sheet['O64'].value)*100,'MLT Norte': (sheet['O62'].value)*100,'MLT Nordeste': (sheet['O63'].value)*100,'EAR Sudeste': (sheet['R65'].value)*100,'EAR Sul': (sheet['R64'].value)*100,'EAR Norte': (sheet['R62'].value)*100,'EAR Nordeste': (sheet['R63'].value)*100, 'Carga Sudeste': sheet['O40'].value,'Carga Sul': sheet['O48'].value, 'Carga Norte': sheet['O24'].value, 'Carga Nordeste': sheet['O32'].value}
        # Transforma para dataframe
        df_ipdo_dia = pd.DataFrame([dados])
        # Concatena os IPDO's
        df_ipdo = pd.concat([df_ipdo,df_ipdo_dia])
    return df_ipdo

def plot_ipdo():
    # Verifica se a coluna 'Data' existe no DataFrame
    if 'Data' not in df_ipdo.columns:
        raise ValueError("O DataFrame precisa ter uma coluna 'Data'")
    
    # Converte a coluna 'Data' para o tipo datetime
    df_ipdo['Data'] = pd.to_datetime(df_ipdo['Data'], format='%d-%m-%Y')
    
    # Seleciona todas as colunas exceto 'Data'
    columns = [col for col in df_ipdo.columns if col != 'Data']
    
    # Cria a pasta 'Gráficos' se não existir
    output_folder = 'Gráficos'
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # Número de colunas para organizar os subplots (ajustável)
    n_cols = 4
    n_rows = (len(columns) + n_cols - 1) // n_cols  # Calcula o número de linhas
    
    # Função auxiliar para criar gráficos de linha
    def plot_graph(df, column_name, title, ax):
        sns.lineplot(data=df, x='Data', y=column_name, ax=ax)
        ax.set_title(title, fontsize=14)
        ax.set_xlabel('Data', fontsize=12)
        ax.set_ylabel(column_name, fontsize=12)
        ax.tick_params(axis='x', rotation=45)
        # Define o formato de data para mostrar apenas dia e mês
        ax.xaxis.set_major_formatter(plt.matplotlib.dates.DateFormatter('%d-%m'))
    
    # Criando a figura com subplots para exibir todos juntos
    fig, axes = plt.subplots(nrows=n_rows, ncols=n_cols, figsize=(20, n_rows * 4), constrained_layout=True)

    # Se houver apenas uma linha, corrigimos a estrutura do array 'axes'
    if n_rows == 1:
        axes = [axes]
    
    # Criando cada gráfico e salvando individualmente em formato JPEG
    for i, col in enumerate(columns):
        ax = axes[i//n_cols][i%n_cols] if n_rows > 1 else axes[i%n_cols]
        plot_graph(df_ipdo, col, col, ax)

        # Salva cada gráfico individualmente no formato JPEG
        fig_individual, ax_individual = plt.subplots(figsize=(8, 6))
        plot_graph(df_ipdo, col, col, ax_individual)

        # Ajusta o layout para evitar o corte da parte inferior (datas)
        fig_individual.tight_layout()

        # Salva o arquivo no formato JPEG
        file_path = os.path.join(output_folder, f"{col}.jpeg")
        fig_individual.savefig(file_path, format='jpeg', dpi=300)
        plt.close(fig_individual)  # Fecha o gráfico individual após salvar
    
    # Removendo gráficos vazios se houver menos subplots que espaços disponíveis
    for j in range(i + 1, n_rows * n_cols):
        fig.delaxes(axes[j//n_cols][j%n_cols] if n_rows > 1 else axes[j%n_cols])

    # Salva o gráfico com todos os subplots em um único arquivo JPEG
    full_plot_path = os.path.join(output_folder, "todos_os_graficos.jpeg")
    fig.savefig(full_plot_path, format='jpeg', dpi=300)
    
    # Exibir os gráficos no subplot
    plt.show()

def create_xlsx():
    # Limpar planilha antiga 
    excel = openpyxl.load_workbook('ipdo/Preenchimento IPDO.xlsx')
    planilha = excel.active
    planilha.delete_rows(1,planilha.max_row)
    excel.save('ipdo/Preenchimento IPDO.xlsx')

    df_ipdo.to_excel('ipdo/Preenchimento IPDO.xlsx', index=False) 

if __name__ == "__main__":
    config = get_info()

    # Faz o download dos IPDO's faltantes da pasta_ipdo
    print('Fazendo download do IPDO')
    ipdo_download()

    # Define quais datas serão
    datas_ipdo = def_datas()

    # Monta dataframe com os IPDO's baixados
    df_ipdo = create_df()
    
    # Plota gráficos
    print('Plotando gráficos')
    plot_ipdo()

    # Monta a planilha xlsx com o dataframe dos ipdo's
    print('Gerando xlsx')
    create_xlsx()